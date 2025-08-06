from django.contrib import admin
from .models import EmployeeProfile, Attendance, EmployeeSalary
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl


@admin.register(EmployeeProfile)
class EmployeeProfileAdmin(admin.ModelAdmin):
    list_display = ['user', 'employee_id', 'position', 'qr_code_tag']
    readonly_fields = ['qr_code_tag']
    search_fields = ['user__username', 'employee_id', 'position']
    list_filter = ['position']
    

admin.site.register(EmployeeSalary)

class AttendanceAdmin(admin.ModelAdmin):
    list_display = ('employee', 'date', 'position', 'punch_in', 'punch_out', 'status', 'permission', 'reason')
    list_filter = ('date', 'employee', 'status', 'position')
    search_fields = ('employee__username',)
    actions = ['export_as_excel']  # ✅ Add this line

    def export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        # Define column headers
        headers = ['Employee', 'Date', 'Status']
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Fill in the rows with data
        for row_num, attendance in enumerate(queryset, start=2):
            ws[f'A{row_num}'] = str(attendance.employee)  # Make sure __str__ is defined on employee
            ws[f'B{row_num}'] = attendance.date.strftime('%Y-%m-%d')
            ws[f'C{row_num}'] = attendance.status

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=attendance_records.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "Download selected as Excel"

